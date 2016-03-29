from tkinter import *
from tkinter import ttk
from openpyxl import *
import requests
import sqlite3


class GUI(Frame):

	def __init__(self, master):
		Frame.__init__(self)
		self.using_home = False
		ttk.Style().configure('back_next.TButton', font = (20), ipadx = 10)
		self.home_page()

	def erase_widgets(self):
		'''Destroys all widgets on a page'''
		for widget in self.master.winfo_children():
			widget.destroy()

	def home_page(self):
		'''The Home Page of the application'''
		self.master.title('Mileage Calculator: Home')
		self.erase_widgets()
		ttk.Style().configure('Justify.TButton', justify = 'center')
		button_container = ttk.PanedWindow(orient = 'horizontal')
		button_container.pack( fill = 'both', expand = 1)
		tutorial = ttk.Button(button_container, text = 'Quick\n Overview', style = 'Justify.TButton', command = self.overview)
		button_container.add(tutorial, weight = 1)
		calculator = ttk.Button(button_container, text = 'Mileage\n Calculator', style = 'Justify.TButton', command = self.options_page)
		button_container.add(calculator, weight = 1)
		addresses = ttk.Button(button_container, text = 'Address\n List', style = 'Justify.TButton', command = self.address_list)
		button_container.add(addresses, weight = 1)

	def options_page(self):
		'''Displays various options for the mileage calculator'''
		self.erase_widgets()
		self.master.title('Mileage Calculator: Options Page')
		self.using_home = False
		back_next_container = PanedWindow()
		back_next_container.pack(side = 'bottom', fill = 'x')
		home_page_button = ttk.Button(back_next_container, text = '\nHome Page\n', style = 'back_next.TButton', command = self.home_page)
		back_next_container.add(home_page_button, sticky = 'w')
		export_choice_label = ttk.Label(self.master, text = 'What do you want to be done with the input info and calculated mileage?', font = (20), padding = '0 0 0 15').pack()
		export_choices = ['Write the information to excel', 'Write the information to a text file', 'Show me the calculated mileage; do nothing with the information']
		export_choice_drop_down = ttk.Combobox(self.master, width = 70, state = 'readonly', values = export_choices)
		export_choice_drop_down.pack()

		def display_next(round = False):
			'''Displays the 'Next' button'''
			if round == True:
				self.round = True
			else:
				self.round = False
			self.rounding_yes_button.config(state = 'disabled')
			self.rounding_no_button.config(state = 'disabled')
			next_button = ttk.Button(back_next_container, text = '\nNext\n', style = 'back_next.TButton', command = self.entry_page)
			back_next_container.add(next_button, sticky = 'e')

		def mileage_rounding(home = None):
			'''Asks the user if they want the mileage result rounded'''
			if home == None:
				self.yes_button.config(state = 'disabled')
				self.no_button.config(state = 'disabled')
			else:
				self.home_yes_button.config(state = 'disabled')
				self.home_no_button.config(state = 'disabled')
			if home == 'hide':
				self.show_home = False
			else:
				self.show_home = True
			question = ttk.Label(self.master, text = 'Would you like the mileage results rounded to the nearest mile?', font = (20), padding = '0 50 0 0').pack()
			yes_no_container = ttk.PanedWindow(self.master, orient = 'horizontal', height = 25)
			yes_no_container.pack()
			self.rounding_yes_button = ttk.Button(yes_no_container, text = 'Yes', command = lambda: display_next(True))
			self.rounding_no_button = ttk.Button(yes_no_container, text = 'No', command = display_next)
			yes_no_container.add(self.rounding_yes_button)
			yes_no_container.add(self.rounding_no_button)

		def home_option():
			'''Asks the user if they want home to show up in the file'''
			question = ttk.Label(self.master, text = 'Would you like the home destination to show up in your file?', font = (20), padding = '0 50 0 0').pack()
			yes_no_container = ttk.PanedWindow(self.master, orient = 'horizontal', height = 25)
			yes_no_container.pack()
			self.home_yes_button = ttk.Button(yes_no_container, text = 'Yes', command = lambda: mileage_rounding('show'))
			self.home_no_button = ttk.Button(yes_no_container, text = 'No', command = lambda: mileage_rounding('hide'))
			yes_no_container.add(self.home_yes_button)
			yes_no_container.add(self.home_no_button)

		def set_home_location():
			'''Adds the home location to the database'''
			home = self.set_entry.get()
			Address_DB.add_location('Home', home)
			home_option()

		def ask_home_location():
			'''If the user has no home address, this function gives the user the option to set one.'''
			self.using_home = True
			self.yes_button.config(state = 'disabled')
			self.no_button.config(state = 'disabled')
			if Address_DB.is_valid_location('Home'):
				if self.export_choice == 'Show':
					mileage_rounding()
				else:
					home_option()
			else:
				set_label = ttk.Label(self.master, text = 'You do not have a home location set.  Please enter the address for your home location: ', font = (20), padding = '0 20 0 0').pack()
				entry_OK_container = ttk.PanedWindow(self.master, orient = 'horizontal', height = 25)
				entry_OK_container.pack()
				self.set_entry = ttk.Entry(entry_OK_container, width = 70)
				ok_button = ttk.Button(entry_OK_container, text = 'OK', command = set_home_location)
				entry_OK_container.add(self.set_entry)
				entry_OK_container.add(ok_button)

		def home_location_widgets():
			'''Loads the widgets necessary to ask the user if they want to use the home location'''
			question = ttk.Label(self.master, text = 'Would you like to use a home address? Using a home address automatically adds that address as a location to the beginning and end of every day.', font = (20), padding = '0 50 0 10').pack()
			yes_no_container = ttk.PanedWindow(self.master, orient = 'horizontal', height = 25)
			yes_no_container.pack()
			self.yes_button = ttk.Button(yes_no_container, text = 'Yes', command = ask_home_location)
			self.no_button = ttk.Button(yes_no_container, text = 'No', command = mileage_rounding)
			yes_no_container.add(self.yes_button)
			yes_no_container.add(self.no_button)

		def check_file():
			'''Confirms the existence of the entered file and adds the appropriate widgets'''
			self.file_name = self.file_name_entry.get()
			if self.export_choice == 'Excel':
				try:
					load_workbook(self.file_name)
				except Exception:
					incorrect_file_name = ttk.Label(self.master, text = 'There is no file named ' + self.file_name + ' in the project folder', padding = '0 5 0 0', foreground = 'red').pack()
				else:
					self.file_name_entry.config(state = 'disabled')
					self.file_name_button.config(state = 'disabled')
					home_location_widgets()
			else:
				try:
					open(self.file_name)
				except Exception:
					incorrect_file_name = ttk.Label(self.master, text = 'There is no file named ' + self.file_name + ' in the project folder', padding = '0 5 0 0', foreground = 'red').pack()
				else:
					self.file_name_entry.config(state = 'disabled')
					self.file_name_button.config(state = 'disabled')
					home_location_widgets()

		def read_export_choice(event):
			'''Reads the selection for export_choice_drop_down and adds the appropriate widgets'''
			export_choice = export_choice_drop_down.current()
			if export_choice == 0:
				self.export_choice = 'Excel'
				file_name_request = ttk.Label(self.master, text = 'Please enter the name of the excel file you wish to write to:', font = (20), padding = '0 50 0 15').pack()
				file_name_reminder = ttk.Label(self.master, text = '(Remember to make sure the file is in the program folder and you added the extension .xlsx to the file name)', font = (20), padding = '0 0 0 30').pack()
				entry_button_container = ttk.PanedWindow(self.master, orient = 'horizontal', height = 25)
				entry_button_container.pack()
				self.file_name_entry = ttk.Entry(entry_button_container, width = 70)
				self.file_name_button = ttk.Button(entry_button_container, text = 'OK', padding = '0 0 0 30', command = check_file)
				entry_button_container.add(self.file_name_entry)
				entry_button_container.add(self.file_name_button)
			elif export_choice == 1:
				self.export_choice = 'Text'
				file_name_request = ttk.Label(self.master, text = 'Please enter the name of the text file you wish to write to:', font = (20), padding = '0 50 0 10').pack()
				file_name_reminder = ttk.Label(self.master, text = '(Remember to make sure the file is in the program folder and you added the extension .txt to the file name)', font = (20), padding = '0 0 0 30').pack()
				entry_button_container = ttk.PanedWindow(self.master, orient = 'horizontal', height = 25)
				entry_button_container.pack()
				self.file_name_entry = ttk.Entry(entry_button_container, width = 70)
				self.file_name_button = ttk.Button(entry_button_container, text = 'OK', padding = '0 0 0 30', command = check_file)
				entry_button_container.add(self.file_name_entry)
				entry_button_container.add(self.file_name_button)
			else:
				self.export_choice = 'Show'
				home_location_widgets()

			export_choice_drop_down.config(state = 'disabled')

		export_choice_drop_down.bind('<<ComboboxSelected>>', read_export_choice)

	def entry_page(self):
		'''The page of the application which allows users to enter locations'''
		self.erase_widgets()
		self.master.title('Mileage Calculator: Entry Page')
		back_next_container = PanedWindow(self.master)
		back_next_container.pack(side = 'bottom', fill = 'x')
		options_page_button = ttk.Button(back_next_container, text = '\nOptions Page\n', style = 'back_next.TButton', command = self.options_page)

		def addresses_window():
			window = Toplevel()
			window.title('Mileage Calculator: Address Table')
			vscrollbar = Scrollbar(window, orient = 'vertical')
			vscrollbar.pack(fill = 'y', side = 'right')
			canvas = Canvas(window, yscrollcommand = vscrollbar.set)
			canvas.pack(side = 'top', fill = 'both', expand = True)
			vscrollbar.config(command = canvas.yview)
			interior = Frame(canvas)
			interior_id = canvas.create_window(0, 0, window = interior, anchor = 'nw')

			def configure_interior(event):
				size = (interior.winfo_reqwidth(), interior.winfo_reqheight())
				canvas.config(scrollregion="0 0 %s %s" % size)
				if interior.winfo_reqwidth() != canvas.winfo_width():
					canvas.config(width = interior.winfo_reqwidth())

			interior.bind('<Configure>', configure_interior)

			def configure_canvas(event):
				if interior.winfo_reqwidth() != canvas.winfo_width():
					canvas.itemconfigure(interior_id, width = canvas.winfo_width())

			canvas.bind('<Configure>', configure_canvas)
			addresses = Address_DB.return_address_dict()
			for location, address in sorted(addresses.items()):
				container = PanedWindow(interior)
				container.pack(fill = 'x')
				location_label = ttk.Label(container, text = location + ':', font = (15))
				address_label = ttk.Label(container, text = address, font = (15))
				container.add(location_label, sticky = 'w')
				container.add(address_label, sticky = 'e')

		address_page_button = ttk.Button(back_next_container, text = '\nAddress Table\n', style = 'back_next.TButton', command = addresses_window)
		back_next_container.add(options_page_button, sticky = 'w')
		back_next_container.add(address_page_button, padx = self.master.winfo_width()/2 - 200)
		self.entry_list = list()
		if self.export_choice != 'Show':
			date_label = ttk.Label(self.master, text = 'Date:', font = (20), padding = '5 5 0 0').pack(anchor = 'nw')
			self.date_entry = ttk.Entry(self.master)
			self.date_entry.pack(anchor = 'nw', padx = 5, pady = 10)
		locations_label = ttk.Label(self.master, text = 'Locations:', font = (20), padding = '5 20 0 10').pack(anchor = 'nw')

		def add_location_entry():
			'''Adds a location entry box'''
			if len(self.entry_list) < 13:
				entry_container = PanedWindow(self.master)
				entry_container.pack(anchor = 'nw', padx = 5, pady = 10)
				entry = ttk.Entry(entry_container, width = 40)
				entry_container.add(entry)
				self.entry_list.append(entry)

		def remove_location_entry():
			'''Removes a location entry box'''
			entry = self.entry_list.pop()
			entry.master.destroy()
			entry.destroy()

		for i in range(3):
			add_location_entry()

		add_remove_container = PanedWindow(self.master, height = 25)
		add_remove_container.pack(side = 'bottom', anchor = 'w', padx = 5, pady = 15)
		add_entry_button = ttk.Button(add_remove_container, text = 'Add Location', command = add_location_entry, width = 19)
		remove_entry_button = ttk.Button(add_remove_container, text = 'Remove Location', command = remove_location_entry, width = 19)
		add_remove_container.add(add_entry_button)
		add_remove_container.add(remove_entry_button)

		def check_entries():
			'''Checks that all entries are in the address database'''
			valid_entries = True
			for address_entry, i in zip(address_list, range(len(address_list))):
				location = invalid_location_list[i]
				address = address_entry.get()
				if location != '' and address != '':
					Address_DB.add_location(location, address)
			for entry_widget in self.entry_list:
				entry = entry_widget.get()
				if not Address_DB.is_valid_location(entry):
					if len(entry_widget.master.panes()) > 1:
						entry_widget.master.forget(entry_widget.master.panes()[1])
						entry_widget.master.forget(entry_widget.master.panes()[1])
					valid_entries = False
					invalid_location_list.append(entry)
					no_address_label = ttk.Label(entry_widget.master, text = 'Please enter an address for ' + entry + ':')
					address_entry = ttk.Entry(entry_widget.master, width = 40)
					address_list.append(address_entry)
					entry_widget.master.add(no_address_label)
					entry_widget.master.add(address_entry)
			if valid_entries:
				self.result_page()

		address_list = list()
		invalid_location_list = list()
		submit_button = ttk.Button(back_next_container, text = '\nSubmit\n', style = 'back_next.TButton', command = check_entries)
		back_next_container.add(submit_button, sticky = 'e')

	def result_page(self):
		'''This page shows the mileage results and confirms if any files have been written to'''
		self.entry_list = list(map(lambda entry: entry.get(), self.entry_list))
		self.master.title('Mileage Calculator: Results')
		if self.using_home == True:
			self.entry_list.insert(0, 'Home')
			self.entry_list.append('Home')
		location_list = self.entry_list[:]
		if self.using_home == True and self.show_home == False:
			del location_list[0]
			location_list.pop()
		locations = entry_list_to_string(location_list)
		addresses = locations_to_addresses(self.entry_list)
		address_string = entry_list_to_string(addresses)
		mileage = Mapquest.calculate_mileage(addresses)
		if self.round == True:
			mileage = round(mileage)
		if self.export_choice == 'Excel':
			date = self.date_entry.get()
			self.erase_widgets()
			workbook = Excel_File(self.file_name)
			workbook.write(date, 'date')
			workbook.write(locations, 'travel_destinations')
			workbook.write(mileage, 'mileage')
			try:
				workbook.wb.save(workbook.filename)
			except PermissionError:
				fail_success_label = ttk.Label(self.master, text = 'Information not written, please close the excel file you want written to.', font = (20))
			else:
				fail_success_label = ttk.Label(self.master, text = 'Information successfully written', font = (20))
			fail_success_label.pack()
		elif self.export_choice == 'Text':
			date = self.date_entry.get()
			tfile = open(self.file_name, 'a')
			tfile.write('\nDate: ' + date + '   ' + 'Locations: ' + locations + '    ' +  'Miles: ' + str(mileage))
			tfile.close()
			self.erase_widgets()
			success_label = ttk.Label(self.master, text = 'Information successfully written.  If not, close the text file you want written to.', font = (20)).pack()
		else:
			self.erase_widgets()
		mileage_label = ttk.Label(self.master, text = 'Miles Traveled: ' + str(mileage), font = (20)).pack()
		locations_label = ttk.Label(self.master, text ='Locations: ' + locations, font = (20), wraplength = self.master.winfo_width()).pack()
		addresses_label = ttk.Label(self.master, text = 'Addresses: ' + address_string, font = (20), wraplength = self.master.winfo_width()).pack()
		back_next_container = PanedWindow(self.master)
		back_next_container.pack(side = 'bottom', fill = 'x')
		entry_page_button = ttk.Button(back_next_container, text = '\nEntry Page\n', style = 'back_next.TButton', command = self.entry_page)
		home_page_button = ttk.Button(back_next_container, text = '\nHome Page\n', style = 'back_next.TButton', command = self.home_page)
		back_next_container.add(entry_page_button, sticky = 'w')
		back_next_container.add(home_page_button, sticky = 'e')

	def address_list(self):
		self.master.title('Mileage Calculator: Addresses')
		self.erase_widgets()
		vscrollbar = Scrollbar(self.master, orient = 'vertical')
		vscrollbar.pack(fill = 'y', side = 'right')
		canvas = Canvas(self.master, yscrollcommand = vscrollbar.set)
		canvas.pack(side = 'top', fill = 'both', expand = True)
		vscrollbar.config(command = canvas.yview)
		button_container = PanedWindow(self.master)
		button_container.pack(side = 'bottom')
		home = ttk.Button(button_container, text = '\nHome Page\n', style = 'back_next.TButton', command = self.home_page)
		add = ttk.Button(button_container, text = '\nAdd Location\n', style = 'back_next.TButton', command = self.add_address)
		remove = ttk.Button(button_container, text = '\nRemove Location\n', style = 'back_next.TButton', command = self.remove_address)
		change = ttk.Button(button_container, text = '\nChange Address\n', style = 'back_next.TButton', command = self.change_address)
		button_container.add(home)
		button_container.add(add)
		button_container.add(remove)
		button_container.add(change)
		interior = Frame(canvas)
		interior_id = canvas.create_window(0, 0, window = interior, anchor = 'nw')

		def configure_interior(event):
			size = (interior.winfo_reqwidth(), interior.winfo_reqheight())
			canvas.config(scrollregion="0 0 %s %s" % size)
			if interior.winfo_reqwidth() != canvas.winfo_width():
				canvas.config(width = interior.winfo_reqwidth())

		interior.bind('<Configure>', configure_interior)

		def configure_canvas(event):
			if interior.winfo_reqwidth() != canvas.winfo_width():
				canvas.itemconfigure(interior_id, width = canvas.winfo_width())

		canvas.bind('<Configure>', configure_canvas)
		addresses = Address_DB.return_address_dict()
		for location, address in sorted(addresses.items()):
			container = PanedWindow(interior)
			container.pack(fill = 'x')
			location_label = ttk.Label(container, text = location + ':', font = (15))
			address_label = ttk.Label(container, text = address, font = (15))
			container.add(location_label, sticky = 'w')
			container.add(address_label, sticky = 'e')

	def add_address(self):
		self.erase_widgets()
		back_next_container = PanedWindow(self.master)
		back_next_container.pack(side = 'bottom', fill = 'x')
		address_page = ttk.Button(back_next_container, text = '\nAddress List\n', style = 'back_next.TButton', command = self.address_list)
		back_next_container.add(address_page, sticky = 'w')
		location_request = ttk.Label(self.master, text = 'Enter the name of the location you would like to add:', font = (20), padding = '0 5 0 10').pack()
		location_entry = ttk.Entry(self.master, width = 70)
		location_entry.pack()
		address_request = ttk.Label(self.master, text = 'Enter the address of this location:', font = (20), padding = '0 50 0 10').pack()
		address_entry = ttk.Entry(self.master, width = 70)
		address_entry.pack()

		def submit():
			location = location_entry.get()
			address = address_entry.get()
			Address_DB.add_location(location, address)
			self.address_list()

		submit = ttk.Button(back_next_container, text = '\nSubmit\n', style = 'back_next.TButton', command = submit)
		back_next_container.add(submit, sticky = 'e')

	def remove_address(self):
		self.erase_widgets()
		back_next_container = PanedWindow(self.master)
		back_next_container.pack(side = 'bottom', fill = 'x')
		address_page = ttk.Button(back_next_container, text = '\nAddress List\n', style = 'back_next.TButton', command = self.address_list)
		back_next_container.add(address_page, sticky = 'w')
		location_request = ttk.Label(self.master, text = 'Enter the name of the location you would like to remove:', font = (20), padding = '0 5 0 10').pack()
		location_entry = ttk.Entry(self.master, width = 70)
		location_entry.pack()

		def submit():
			location = location_entry.get()
			if Address_DB.is_valid_location(location):
				Address_DB.remove_location(location)
				self.address_list()
			else:
				wrong_location = ttk.Label(self.master, text = 'There is no location with the name ' + location + ' in the database.', foreground = 'red').pack()

		submit = ttk.Button(back_next_container, text = '\nSubmit\n', style = 'back_next.TButton', command = submit)
		back_next_container.add(submit, sticky = 'e')

	def change_address(self):
		self.erase_widgets()
		back_next_container = PanedWindow(self.master)
		back_next_container.pack(side = 'bottom', fill = 'x')
		address_page = ttk.Button(back_next_container, text = '\nAddress List\n', style = 'back_next.TButton', command = self.address_list)
		back_next_container.add(address_page, sticky = 'w')

		def submit(new_location, new_address):
			Address_DB.change_location_address(location_entry.get(), new_address)
			Address_DB.change_location_name(new_location, new_address)
			self.address_list()

		def update():
			location = location_entry.get()
			address = Address_DB.return_address(location)
			if Address_DB.is_valid_location(location):
				location_entry.config(state = 'disabled')
				label_container = PanedWindow(self.master)
				label_container.pack()
				location_request = ttk.Label(label_container, text = 'Location:', font = (20), padding = '0 50 350 10')
				address_request = ttk.Label(label_container, text = 'Address:', font = (20), padding = '0 50 0 10')
				label_container.add(location_request, sticky = 'w')
				label_container.add(address_request, sticky = 'w')
				entry_container = PanedWindow(self.master)
				entry_container.pack()
				new_location_entry = Entry(entry_container, width = 70)
				new_address_entry = Entry(entry_container, width = 70)
				new_location_entry.insert(0, location_entry.get())
				new_address_entry.insert(0, Address_DB.return_address(location_entry.get()))
				entry_container.add(new_location_entry)
				entry_container.add(new_address_entry)
				submit_button = ttk.Button(back_next_container, text = '\nSubmit\n', style = 'back_next.TButton', command = lambda: submit(new_location_entry.get(), new_address_entry.get()))
				back_next_container.add(submit_button, sticky = 'e')
			else:
				wrong_location = ttk.Label(self.master, text = 'There is no location with the name ' + location + ' in the database.', foreground = 'red').pack()


		location_request = ttk.Label(self.master, text = 'Enter the name of the location you would like to change:', font = (20), padding = '0 5 0 10').pack()
		location_entry_container = PanedWindow(self.master, height = 25)
		location_entry_container.pack()
		location_entry = ttk.Entry(location_entry_container, width = 70)
		ok_button = ttk.Button(location_entry_container, text = 'OK', command = update)
		location_entry_container.add(location_entry)
		location_entry_container.add(ok_button)

	def overview(self):
		self.erase_widgets()
		self.master.title('Mileage Calculator: Overview')
		text = ttk.Label(self.master, text = ('The purpose of this program is to allow small businesses to track company driving mileage, and therefore be able to calculate gas '
						'expenditure for tax purposes.  The program allows users to enter the name of a location, instead of the address(ex. Panera-Independence instead of '
						'6700 Rockside Rd, Independence, OH 44131).  Another useful feature is the option of using a home location.  With this option, the program will automatically '
						'add your home or business address to the beginning and end of your trip.  Additionally, the program can automatically add the date, locations, and mileage '
						'of each day to an excel or text file for record keeping purposes.'	), wraplength = self.master.winfo_width(), font = (20), padding = '0 0 0 50').pack()
		contact = ttk.Label(self.master, text = 'If there are any issues, questions, or comments with or about this program please feel free to email me at clayw.james@gmail.com', font = (20)).pack()
		home_errors_container = PanedWindow(self.master)
		home_errors_container.pack(fill = 'x', side = 'bottom')
		home_page = ttk.Button(home_errors_container, text = '\nHome Page\n', style = 'back_next.TButton', command = self.home_page)
		errors_page = ttk.Button(home_errors_container, text = '\nError Solutions\n', style = 'back_next.TButton', command = self.errors_page)
		home_errors_container.add(home_page, sticky = 'w')
		home_errors_container.add(errors_page, sticky = 'e')

	def errors_page(self):
		self.erase_widgets()
		self.master.title('Mileage Calculator: Error Solutions')
		title = ttk.Label(self.master, text = 'Possible solutions to errors:', font = (20), padding = '0 0 0 50').pack()
		error1 = ttk.Label(self.master, text = 'The mileage results are wrong: Go to the Address List and make sure the addresses are correct.').pack(pady = 20)
		error2 = ttk.Label(self.master, text = ('The output file is not being written to: Make sure you are entering the correct file name with the correct file extension'
							'(.txt for text and .xlsx for excel).  Also make sure that the file is in the program folder.  Finally, always close the file you want written to '
							'before you use the program.'), wraplength = self.master.winfo_width()).pack(pady = 20)
		error3 = ttk.Label(self.master, text = "I can't copy and paste: to copy and paste into entry boxes use control/command + c and control/command + v, respectively").pack(pady = 20)
		back = ttk.Button(self.master, text = '\nOverview\n', style = 'back_next.TButton', command = self.overview).pack(side = 'bottom', anchor = 'w')


def entry_list_to_string(entry_list):
    entry_string = ''
    for entry,i in zip(entry_list, range(len(entry_list))):
        if i == 0:
            entry_string = entry
        else:
            entry_string = entry_string + ', ' + entry
    return entry_string


def locations_to_addresses(locations):
    for loc,i in zip(locations,range(len(locations))):
        locations[i] = Address_DB.return_address(loc)
    return locations


class Excel_File:

    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(filename)
        self.ws = self.wb.active

    def first_blank_square(self, column):
        '''Finds the first square in the correct column that hasn't been written to and returns it'''
        row_number = 1
        while True:
            row_number += 1
            row = column + str(row_number)
            if self.ws[row].value == None:
                return row

    def write(self, content, category):
        '''Writes to the file in a the next open location in the correct column'''
        if category == 'date':
            self.ws[self.first_blank_square('A')] = content
        if category == 'travel_destinations':
            self.ws[self.first_blank_square('B')] = content
        if category == 'mileage':
            self.ws[self.first_blank_square('C')] = content


class Mapquest:

    key = 'MkRTKx7DbBySjsya4hnVsQ0bxgQgnbSy'
    mileage = []

    def calculate_mileage(locations):
        '''Calculates the mileage for one day using Mapquest API'''
        for loc,i in zip(locations,range(len(locations))):
            if i == 0:
                locations[0] = '&from=' + loc
            else:
                locations[i] = '&to=' + loc
        url = 'http://www.mapquestapi.com/directions/v2/route?key=' + Mapquest.key
        for loc in locations:
            url = url + loc
        request = requests.get(url)
        response = request.json()
        return response['route']['distance']


class Address_DB:

    def is_valid_location(location):
        con = sqlite3.connect('addresses.db')
        c = con.cursor()
        c.execute('select act_address from addresses where loc_name = ? ', [str(location)])
        address = c.fetchone()
        if address is None:
            return False
        else:
            return True

    def add_location(location, address):
        con = sqlite3.connect('addresses.db')
        c = con.cursor()
        c.execute('insert into addresses values (@0, @1)', (location, address))
        con.commit()
        con.close()

    def remove_location(location):
        con = sqlite3.connect('addresses.db')
        c = con.cursor()
        c.execute('delete from addresses where loc_name = ?', [location])
        con.commit()
        con.close()

    def change_location_address(location, address):
        con = sqlite3.connect('addresses.db')
        c = con.cursor()
        c.execute('update addresses set act_address = @0 where loc_name = @1',(address, location))
        con.commit()
        con.close()

    def change_location_name(location, address):
    	con = sqlite3.connect('addresses.db')
    	c = con.cursor()
    	c.execute('update addresses set loc_name = @0 where act_address = @1',(location, address))
    	con.commit()
    	con.close()

    def return_address(location):
        con = sqlite3.connect('addresses.db')
        c = con.cursor()
        c.execute('select act_address from addresses where loc_name = ?',[location])
        address = c.fetchone()
        return address[0]

    def return_address_dict():
    	con = sqlite3.connect('addresses.db')
    	c = con.cursor()
    	address_dict = dict()
    	c.execute('select * from addresses')
    	a_list = c.fetchall()
    	for entry in a_list:
    		address_dict[entry[0]] = entry[1]
    	return address_dict

con = sqlite3.connect('addresses.db')
c = con.cursor()
c.execute('create table if not exists addresses (loc_name, act_address)')
con.commit()
con.close()

root = Tk()
root.state('zoomed')
gui = GUI(master = root)
gui.mainloop()
